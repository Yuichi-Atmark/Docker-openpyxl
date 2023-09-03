import openpyxl as px
import json
import csv

# 出力データ
json_list = []
# 見出し行退避用
cap = []
# csvrow = 'csvrow'
#
# Excelファイルの置き場所を指定
#
inputdir = '/root/input'
outputdir = '/root/output'

def setData(colData):
    if isinstance(colData, (int, float)):
        return colData
    else:
        return str(colData)

def main():
    #
    # Excelファイルを開く
    #
    wb = px.load_workbook( inputdir + '/RESULT.xlsx')
    print(type(wb))

    #
    # ワークシートを開く
    #
    ws = wb.worksheets[0]

    #
    # セルを読み込む
    #

    # 1～12行
    for rows in range(1, 12):
        rowdata = ""

        # 1行目のとき列名として保存
        if rows == 1:
            # 列名を保存
            for cols in range(1, 8):
                cap.append( ws.cell(rows, cols).value )
        # 2行目以降はデータとして扱う
        else:
            # print(ws.cell(rows, cols).number_format + ': ' + str(ws.cell(rows, cols).value) )
            # rowdata = rowdata + cap[cols - 1] + ': ' + str(ws.cell(rows, cols).value) + ','
            # for convcol in range(1, 8):
            #    if isinstance(ws.cell(rows, convcol).value, (int, float)):
            #        print ("{0}列目のセルの書式は 数字項目 です".format(convcol))

            rowdata = {
                "csvrow": {
                    cap[i - 1]: setData(ws.cell(rows, i).value) for i in range(1, 8)
                }
            }
            
        # 行データ書き出し
        if rows > 1:
            json_list.append(rowdata)

    p2j_data  = json.dumps(json_list, indent=2)
    print(json.dumps(json_list, indent=2))

    #
    # JSONデータ書き出し
    #
    with open(outputdir + '/output.json', 'w') as f:
        json.dump(json_list, f)

    wb.close

if __name__ == "__main__":
    main()
